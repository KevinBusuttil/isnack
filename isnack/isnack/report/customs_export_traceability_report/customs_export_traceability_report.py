# Copyright (c) 2026, Busuttil Technologies Limited and contributors
# For license information, please see license.txt

import base64
import json
import os
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import getdate, nowdate, now_datetime


def execute(filters=None):
	filters = frappe._dict(filters or {})
	validate_filters(filters)
	columns = get_columns()
	data = get_data(filters)
	return columns, data


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_filters(filters):
	for fld in ("company", "from_date", "to_date"):
		if not filters.get(fld):
			frappe.throw(_("Filter {0} is required").format(fld))
	if getdate(filters.from_date) > getdate(filters.to_date):
		frappe.throw(_("From Date must be on or before To Date"))


# ---------------------------------------------------------------------------
# Columns
# ---------------------------------------------------------------------------

def get_columns():
	return [
		# A. Sales Invoice context
		{"label": _("Company"), "fieldname": "company", "fieldtype": "Link", "options": "Company", "width": 120},
		{"label": _("Sales Invoice"), "fieldname": "sales_invoice", "fieldtype": "Link", "options": "Sales Invoice", "width": 150},
		{"label": _("Posting Date"), "fieldname": "posting_date", "fieldtype": "Date", "width": 100},
		{"label": _("Customer"), "fieldname": "customer", "fieldtype": "Link", "options": "Customer", "width": 120},
		{"label": _("Customer Name"), "fieldname": "customer_name", "fieldtype": "Data", "width": 160},
		{"label": _("Currency"), "fieldname": "currency", "fieldtype": "Link", "options": "Currency", "width": 80},
		{"label": _("Row #"), "fieldname": "si_item_idx", "fieldtype": "Int", "width": 60},
		{"label": _("FG Item Code"), "fieldname": "fg_item_code", "fieldtype": "Link", "options": "Item", "width": 140},
		{"label": _("FG Item Name"), "fieldname": "fg_item_name", "fieldtype": "Data", "width": 160},
		{"label": _("FG Description"), "fieldname": "fg_description", "fieldtype": "Data", "width": 180},
		{"label": _("Sales Qty"), "fieldname": "sales_qty", "fieldtype": "Float", "width": 90},
		{"label": _("Sales UOM"), "fieldname": "sales_uom", "fieldtype": "Link", "options": "UOM", "width": 80},
		{"label": _("Stock Qty"), "fieldname": "stock_qty", "fieldtype": "Float", "width": 90},
		{"label": _("Item Group"), "fieldname": "item_group", "fieldtype": "Link", "options": "Item Group", "width": 120},
		# B. Finished good traceability
		{"label": _("FG Batch No"), "fieldname": "fg_batch_no", "fieldtype": "Link", "options": "Batch", "width": 130},
		{"label": _("Work Order"), "fieldname": "work_order", "fieldtype": "Link", "options": "Work Order", "width": 150},
		{"label": _("WO Item"), "fieldname": "wo_item", "fieldtype": "Link", "options": "Item", "width": 130},
		{"label": _("WO Qty"), "fieldname": "wo_qty", "fieldtype": "Float", "width": 80},
		{"label": _("Manufacturing Date"), "fieldname": "manufacturing_date", "fieldtype": "Date", "width": 110},
		{"label": _("Manufacture Entry"), "fieldname": "manufacture_entry", "fieldtype": "Link", "options": "Stock Entry", "width": 150},
		# C. Raw material consumption
		{"label": _("RM Item Code"), "fieldname": "rm_item_code", "fieldtype": "Link", "options": "Item", "width": 140},
		{"label": _("RM Item Name"), "fieldname": "rm_item_name", "fieldtype": "Data", "width": 160},
		{"label": _("RM Description"), "fieldname": "rm_description", "fieldtype": "Data", "width": 180},
		{"label": _("RM UOM"), "fieldname": "rm_uom", "fieldtype": "Link", "options": "UOM", "width": 80},
		{"label": _("Consumed Qty"), "fieldname": "consumed_qty", "fieldtype": "Float", "width": 100},
		{"label": _("RM Batch No"), "fieldname": "rm_batch_no", "fieldtype": "Link", "options": "Batch", "width": 130},
		# D. Purchase / customs traceability
		{"label": _("Purchase Receipt"), "fieldname": "purchase_receipt", "fieldtype": "Link", "options": "Purchase Receipt", "width": 150},
		{"label": _("PR Date"), "fieldname": "purchase_receipt_date", "fieldtype": "Date", "width": 100},
		{"label": _("Supplier Name"), "fieldname": "supplier_name", "fieldtype": "Data", "width": 160},
		{"label": _("PR Qty"), "fieldname": "pr_qty", "fieldtype": "Float", "width": 100},
		{"label": _("Balance Stock"), "fieldname": "balance_stock", "fieldtype": "Float", "width": 110},
		{"label": _("Customs Document No"), "fieldname": "customs_document_no", "fieldtype": "Data", "width": 180},
	]


# ---------------------------------------------------------------------------
# Data assembly
# ---------------------------------------------------------------------------

def get_data(filters):
	# Step 1: Fetch SI items
	si_items = _fetch_si_items(filters)
	if not si_items:
		return []

	# Step 2: Resolve FG batches per SI item (direct + via bundle)
	bundle_names = [r.serial_and_batch_bundle for r in si_items if r.serial_and_batch_bundle and not r.batch_no]
	bundle_entries = _fetch_bundle_entries(bundle_names)  # {bundle_name: [{batch_no, qty}]}

	# Step 3: Collect all FG (item_code, batch_no) pairs to find Work Orders
	fg_batch_pairs = set()
	for row in si_items:
		if row.batch_no:
			fg_batch_pairs.add((row.fg_item_code, row.batch_no))
		elif row.serial_and_batch_bundle and bundle_entries.get(row.serial_and_batch_bundle):
			for be in bundle_entries[row.serial_and_batch_bundle]:
				if be["batch_no"]:
					fg_batch_pairs.add((row.fg_item_code, be["batch_no"]))

	# Step 4: Resolve Work Orders for FG batches
	wo_map = _fetch_manufacture_entries(fg_batch_pairs)  # {(item_code, batch_no): [{work_order, ...}]}

	# Step 5: Collect all work orders → resolve consumed raw materials
	all_work_orders = set()
	for entries in wo_map.values():
		for e in entries:
			if e.get("work_order"):
				all_work_orders.add(e["work_order"])

	rm_map = _fetch_rm_consumption(all_work_orders, filters)  # {work_order: [rm_dict]}

	# Step 6: Collect all PR names → resolve PR details (incl. customs doc no)
	pr_names = set()
	for rm_list in rm_map.values():
		for rm in rm_list:
			if rm.get("purchase_receipt"):
				pr_names.add(rm["purchase_receipt"])
	pr_details = _fetch_pr_details(pr_names)  # {pr_name: {...}}

	# Step 6b: Fetch PR item quantities for (pr_name, item_code, batch_no) combos
	pr_batch_keys = set()
	for rm_list in rm_map.values():
		for rm in rm_list:
			if rm.get("purchase_receipt") and rm.get("item_code") and rm.get("batch_no"):
				pr_batch_keys.add((rm["purchase_receipt"], rm["item_code"], rm["batch_no"]))
	pr_item_qty_map = _fetch_pr_item_qty(pr_batch_keys)  # {(pr_name, item_code, batch_no): qty}

	# Step 6c: Fetch balance stock for (item_code, batch_no) pairs
	batch_item_pairs = set()
	for rm_list in rm_map.values():
		for rm in rm_list:
			if rm.get("item_code") and rm.get("batch_no"):
				batch_item_pairs.add((rm["item_code"], rm["batch_no"]))
	batch_balance_map = _fetch_batch_balance(batch_item_pairs)  # {(item_code, batch_no): qty}

	# Step 7: Assemble output rows
	rows = []
	for si_row in si_items:
		# Build list of (fg_batch_no, batch_qty_fraction) for this SI item
		fg_batches = _resolve_fg_batches(si_row, bundle_entries)

		for fg_batch_no in fg_batches:
			wo_entries = wo_map.get((si_row.fg_item_code, fg_batch_no)) or [{}]
			for wo_entry in wo_entries:
				work_order = wo_entry.get("work_order")
				rm_list = rm_map.get(work_order) or [{}] if work_order else [{}]
				for rm in rm_list:
					pr_name = rm.get("purchase_receipt")
					pr = pr_details.get(pr_name) or {} if pr_name else {}

					row = frappe._dict(
						# A
						company=si_row.company,
						sales_invoice=si_row.sales_invoice,
						posting_date=si_row.posting_date,
						customer=si_row.customer,
						customer_name=si_row.customer_name,
						currency=si_row.currency,
						si_item_idx=si_row.idx,
						fg_item_code=si_row.fg_item_code,
						fg_item_name=si_row.fg_item_name,
						fg_description=si_row.fg_description,
						sales_qty=si_row.qty,
						sales_uom=si_row.uom,
						stock_qty=si_row.stock_qty,
						item_group=si_row.item_group,
						# B
						fg_batch_no=fg_batch_no or None,
						work_order=work_order,
						wo_item=wo_entry.get("wo_item"),
						wo_qty=wo_entry.get("wo_qty"),
						manufacturing_date=wo_entry.get("manufacturing_date"),
						manufacture_entry=wo_entry.get("stock_entry"),
						# C
						rm_item_code=rm.get("item_code"),
						rm_item_name=rm.get("item_name"),
						rm_description=rm.get("description"),
						rm_uom=rm.get("stock_uom"),
						consumed_qty=rm.get("qty"),
						rm_batch_no=rm.get("batch_no"),
						# D
						purchase_receipt=pr_name,
						purchase_receipt_date=pr.get("posting_date"),
						supplier=pr.get("supplier"),
						supplier_name=pr.get("supplier_name"),
						pr_qty=pr_item_qty_map.get((pr_name, rm.get("item_code"), rm.get("batch_no")), 0) or None,
						balance_stock=batch_balance_map.get((rm.get("item_code"), rm.get("batch_no")), 0) or None,
						customs_document_no=pr.get("custom_customs_document_no"),
					)

					if passes_post_filters(row, filters):
						rows.append(row)

	return rows


# ---------------------------------------------------------------------------
# Post-filter (for filters that can't be pushed to initial SQL)
# ---------------------------------------------------------------------------

def passes_post_filters(row, filters):
	if filters.get("work_order") and row.get("work_order") != filters.work_order:
		return False
	if filters.get("purchase_receipt") and row.get("purchase_receipt") != filters.purchase_receipt:
		return False
	if filters.get("customs_document_no"):
		cdn = (row.get("customs_document_no") or "").lower()
		if filters.customs_document_no.lower() not in cdn:
			return False
	return True


# ---------------------------------------------------------------------------
# Helper: resolve FG batch list for a single SI item row
# ---------------------------------------------------------------------------

def _resolve_fg_batches(si_row, bundle_entries):
	"""Return list of fg_batch_no strings (may be empty string / None if unknown)."""
	if si_row.batch_no:
		return [si_row.batch_no]
	if si_row.serial_and_batch_bundle:
		entries = bundle_entries.get(si_row.serial_and_batch_bundle) or []
		batches = [e["batch_no"] for e in entries if e.get("batch_no")]
		if batches:
			return batches
	return [None]


# ---------------------------------------------------------------------------
# Step 1: Fetch Sales Invoice Items
# ---------------------------------------------------------------------------

def _fetch_si_items(filters):
	conditions = [
		"si.docstatus = 1",
		"si.company = %(company)s",
		"si.posting_date BETWEEN %(from_date)s AND %(to_date)s",
	]
	values = {
		"company": filters.company,
		"from_date": filters.from_date,
		"to_date": filters.to_date,
	}

	if filters.get("sales_invoice"):
		conditions.append("si.name = %(sales_invoice)s")
		values["sales_invoice"] = filters.sales_invoice

	if filters.get("customer"):
		conditions.append("si.customer = %(customer)s")
		values["customer"] = filters.customer

	if filters.get("item_code"):
		conditions.append("sii.item_code = %(item_code)s")
		values["item_code"] = filters.item_code

	if filters.get("item_group"):
		conditions.append("sii.item_group = %(item_group)s")
		values["item_group"] = filters.item_group

	# Batch filter: either directly on sii.batch_no or via bundle
	if filters.get("batch_no"):
		conditions.append("""(
			sii.batch_no = %(batch_no)s
			OR EXISTS (
				SELECT 1 FROM `tabSerial and Batch Entry` sbe
				WHERE sbe.parent = sii.serial_and_batch_bundle
				AND sbe.batch_no = %(batch_no)s
			)
		)""")
		values["batch_no"] = filters.batch_no

	where_clause = " AND ".join(conditions)

	return frappe.db.sql(
		f"""
		SELECT
			si.name          AS sales_invoice,
			si.company,
			si.posting_date,
			si.customer,
			si.customer_name,
			si.currency,
			sii.idx,
			sii.item_code    AS fg_item_code,
			sii.item_name    AS fg_item_name,
			sii.description  AS fg_description,
			sii.qty,
			sii.uom,
			sii.stock_qty,
			sii.item_group,
			sii.batch_no,
			sii.serial_and_batch_bundle
		FROM `tabSales Invoice Item` sii
		JOIN `tabSales Invoice` si ON si.name = sii.parent
		WHERE {where_clause}
		ORDER BY si.posting_date, si.name, sii.idx
		""",
		values,
		as_dict=True,
	)


# ---------------------------------------------------------------------------
# Step 2: Batch-fetch Serial and Batch Bundle entries
# ---------------------------------------------------------------------------

def _fetch_bundle_entries(bundle_names):
	"""Return {bundle_name: [{batch_no, qty}]}"""
	if not bundle_names:
		return {}
	placeholders = ", ".join(["%s"] * len(bundle_names))
	rows = frappe.db.sql(
		f"""
		SELECT parent, batch_no, qty
		FROM `tabSerial and Batch Entry`
		WHERE parent IN ({placeholders})
		AND batch_no IS NOT NULL AND batch_no != ''
		""",
		tuple(bundle_names),
		as_dict=True,
	)
	result = {}
	for r in rows:
		result.setdefault(r.parent, []).append({"batch_no": r.batch_no, "qty": r.qty})
	return result


# ---------------------------------------------------------------------------
# Step 3: Resolve Work Orders from FG (item_code, batch_no) pairs
# ---------------------------------------------------------------------------

def _fetch_manufacture_entries(fg_batch_pairs):
	"""
	Return {(item_code, batch_no): [{work_order, stock_entry, manufacturing_date, wo_item, wo_qty}]}

	Strategy:
	  a. Direct batch_no on Stock Entry Detail (is_finished_item=1)
	  b. Via serial_and_batch_bundle on Stock Entry Detail → Serial and Batch Entry
	"""
	if not fg_batch_pairs:
		return {}

	batch_nos = list({b for _, b in fg_batch_pairs if b})
	if not batch_nos:
		return {}

	placeholders = ", ".join(["%s"] * len(batch_nos))

	# Strategy a: direct batch_no on the finished-item detail row
	direct_rows = frappe.db.sql(
		f"""
		SELECT
			se.name          AS stock_entry,
			se.work_order,
			se.posting_date  AS manufacturing_date,
			sed.item_code,
			sed.batch_no
		FROM `tabStock Entry Detail` sed
		JOIN `tabStock Entry` se ON se.name = sed.parent
		WHERE se.purpose = 'Manufacture'
		  AND se.docstatus = 1
		  AND se.work_order IS NOT NULL AND se.work_order != ''
		  AND sed.is_finished_item = 1
		  AND sed.batch_no IN ({placeholders})
		""",
		tuple(batch_nos),
		as_dict=True,
	)

	# Strategy b: batch_no via serial_and_batch_bundle on finished-item detail row
	bundle_rows = frappe.db.sql(
		f"""
		SELECT
			se.name          AS stock_entry,
			se.work_order,
			se.posting_date  AS manufacturing_date,
			sed.item_code,
			sbe.batch_no
		FROM `tabStock Entry Detail` sed
		JOIN `tabStock Entry` se ON se.name = sed.parent
		JOIN `tabSerial and Batch Entry` sbe ON sbe.parent = sed.serial_and_batch_bundle
		WHERE se.purpose = 'Manufacture'
		  AND se.docstatus = 1
		  AND se.work_order IS NOT NULL AND se.work_order != ''
		  AND sed.is_finished_item = 1
		  AND sed.serial_and_batch_bundle IS NOT NULL
		  AND sbe.batch_no IN ({placeholders})
		""",
		tuple(batch_nos),
		as_dict=True,
	)

	all_se_rows = direct_rows + bundle_rows

	# Bulk fetch Work Order details
	wo_names = list({r.work_order for r in all_se_rows if r.work_order})
	wo_details = {}
	if wo_names:
		wo_placeholders = ", ".join(["%s"] * len(wo_names))
		for wo in frappe.db.sql(
			f"""
			SELECT name, production_item, qty, actual_start_date
			FROM `tabWork Order`
			WHERE name IN ({wo_placeholders})
			""",
			tuple(wo_names),
			as_dict=True,
		):
			wo_details[wo.name] = wo

	result = {}
	seen = set()
	for r in all_se_rows:
		key = (r.item_code, r.batch_no)
		if key not in {(ic, b) for ic, b in fg_batch_pairs}:
			continue
		wo = wo_details.get(r.work_order) or {}
		entry = {
			"work_order": r.work_order,
			"stock_entry": r.stock_entry,
			"manufacturing_date": r.manufacturing_date,
			"wo_item": wo.get("production_item"),
			"wo_qty": wo.get("qty"),
		}
		dedup_key = (key, r.stock_entry)
		if dedup_key not in seen:
			seen.add(dedup_key)
			result.setdefault(key, []).append(entry)

	return result


# ---------------------------------------------------------------------------
# Step 4: Resolve consumed raw materials per Work Order
# ---------------------------------------------------------------------------

def _fetch_rm_consumption(work_orders, filters):
	"""
	Return {work_order: [rm_dict]}

	Each rm_dict has: item_code, item_name, description, stock_uom, qty,
	                  batch_no, purchase_receipt
	"""
	if not work_orders:
		return {}

	wo_list = list(work_orders)
	placeholders = ", ".join(["%s"] * len(wo_list))

	# Fetch consumption rows (is_finished_item=0, has a source warehouse)
	raw_rows = frappe.db.sql(
		f"""
		SELECT
			se.work_order,
			sed.item_code,
			sed.item_name,
			sed.description,
			sed.stock_uom,
			sed.qty,
			sed.batch_no,
			sed.serial_and_batch_bundle,
			sed.reference_purchase_receipt
		FROM `tabStock Entry Detail` sed
		JOIN `tabStock Entry` se ON se.name = sed.parent
		WHERE se.docstatus = 1
		  AND se.work_order IN ({placeholders})
		  AND se.purpose IN ('Manufacture', 'Material Consumption for Manufacture')
		  AND sed.is_finished_item = 0
		  AND sed.s_warehouse IS NOT NULL AND sed.s_warehouse != ''
		""",
		tuple(wo_list),
		as_dict=True,
	)

	# Apply raw_material_item filter early
	if filters.get("raw_material_item"):
		raw_rows = [r for r in raw_rows if r.item_code == filters.raw_material_item]

	# Expand rows that use serial_and_batch_bundle (no direct batch_no)
	bundle_names_rm = [r.serial_and_batch_bundle for r in raw_rows
					   if r.serial_and_batch_bundle and not r.batch_no]
	rm_bundle_entries = _fetch_bundle_entries(bundle_names_rm)

	result = {}
	for r in raw_rows:
		# Build one or more rm dicts (one per batch when using bundle)
		rm_dicts = []
		if r.batch_no:
			rm_dicts.append({
				"item_code": r.item_code,
				"item_name": r.item_name,
				"description": r.description,
				"stock_uom": r.stock_uom,
				"qty": r.qty,
				"batch_no": r.batch_no,
				"purchase_receipt": r.reference_purchase_receipt or None,
			})
		elif r.serial_and_batch_bundle:
			entries = rm_bundle_entries.get(r.serial_and_batch_bundle) or []
			if entries:
				total_bundle_qty = sum(e["qty"] for e in entries if e["qty"])
				for be in entries:
					qty_fraction = (be["qty"] / total_bundle_qty * r.qty) if total_bundle_qty else r.qty
					rm_dicts.append({
						"item_code": r.item_code,
						"item_name": r.item_name,
						"description": r.description,
						"stock_uom": r.stock_uom,
						"qty": qty_fraction,
						"batch_no": be["batch_no"],
						"purchase_receipt": r.reference_purchase_receipt or None,
					})
			else:
				rm_dicts.append({
					"item_code": r.item_code,
					"item_name": r.item_name,
					"description": r.description,
					"stock_uom": r.stock_uom,
					"qty": r.qty,
					"batch_no": None,
					"purchase_receipt": r.reference_purchase_receipt or None,
				})
		else:
			rm_dicts.append({
				"item_code": r.item_code,
				"item_name": r.item_name,
				"description": r.description,
				"stock_uom": r.stock_uom,
				"qty": r.qty,
				"batch_no": None,
				"purchase_receipt": r.reference_purchase_receipt or None,
			})

		# Fallback: if no purchase_receipt from reference_purchase_receipt, try via batch
		for rm in rm_dicts:
			if not rm["purchase_receipt"] and rm.get("batch_no"):
				rm["purchase_receipt"] = _lookup_pr_via_batch(rm["batch_no"], rm["item_code"])

		for rm in rm_dicts:
			result.setdefault(r.work_order, []).append(rm)

	return result


def _lookup_pr_via_batch(batch_no, item_code):
	"""Fallback: find a Purchase Receipt for a given batch_no + item_code."""
	# Try via Purchase Receipt Item directly
	pr = frappe.db.get_value(
		"Purchase Receipt Item",
		{"batch_no": batch_no, "item_code": item_code, "docstatus": 1},
		"parent",
	)
	if pr:
		return pr

	# Try via Serial and Batch Bundle on Purchase Receipt
	result = frappe.db.sql(
		"""
		SELECT sbb.voucher_no
		FROM `tabSerial and Batch Entry` sbe
		JOIN `tabSerial and Batch Bundle` sbb ON sbb.name = sbe.parent
		WHERE sbe.batch_no = %s
		  AND sbb.voucher_type = 'Purchase Receipt'
		  AND sbb.docstatus = 1
		LIMIT 1
		""",
		(batch_no,),
	)
	return result[0][0] if result else None


# ---------------------------------------------------------------------------
# Step 5: Fetch Purchase Receipt details in bulk
# ---------------------------------------------------------------------------

def _fetch_pr_details(pr_names):
	"""Return {pr_name: {posting_date, supplier, supplier_name, custom_customs_document_no}}"""
	if not pr_names:
		return {}
	placeholders = ", ".join(["%s"] * len(pr_names))
	rows = frappe.db.sql(
		f"""
		SELECT
			name,
			posting_date,
			supplier,
			supplier_name,
			custom_customs_document_no
		FROM `tabPurchase Receipt`
		WHERE name IN ({placeholders})
		  AND docstatus = 1
		""",
		tuple(pr_names),
		as_dict=True,
	)
	return {r.name: r for r in rows}


# ---------------------------------------------------------------------------
# Step 6b: Fetch Purchase Receipt Item qty for (pr_name, item_code, batch_no)
# ---------------------------------------------------------------------------

def _fetch_pr_item_qty(pr_batch_keys):
	"""Return {(pr_name, item_code, batch_no): received_qty}"""
	if not pr_batch_keys:
		return {}

	result = {}
	for pr_name, item_code, batch_no in pr_batch_keys:
		if not (pr_name and item_code and batch_no):
			continue

		# Strategy a: direct batch_no on Purchase Receipt Item
		direct_qty = frappe.db.sql(
			"""
			SELECT IFNULL(SUM(pri.qty), 0)
			FROM `tabPurchase Receipt Item` pri
			WHERE pri.parent = %s
			  AND pri.item_code = %s
			  AND pri.batch_no = %s
			  AND pri.docstatus = 1
			""",
			(pr_name, item_code, batch_no),
		)
		qty = direct_qty[0][0] if direct_qty else 0

		# Strategy b: via serial_and_batch_bundle
		if not qty:
			bundle_qty = frappe.db.sql(
				"""
				SELECT IFNULL(SUM(ABS(sbe.qty)), 0)
				FROM `tabPurchase Receipt Item` pri
				JOIN `tabSerial and Batch Entry` sbe ON sbe.parent = pri.serial_and_batch_bundle
				WHERE pri.parent = %s
				  AND pri.item_code = %s
				  AND sbe.batch_no = %s
				  AND pri.docstatus = 1
				  AND pri.serial_and_batch_bundle IS NOT NULL
				""",
				(pr_name, item_code, batch_no),
			)
			qty = bundle_qty[0][0] if bundle_qty else 0

		if qty:
			result[(pr_name, item_code, batch_no)] = qty

	return result


# ---------------------------------------------------------------------------
# Step 6c: Fetch current batch balance stock across all warehouses
# ---------------------------------------------------------------------------

def _fetch_batch_balance(batch_item_pairs):
	"""Return {(item_code, batch_no): total_balance_qty} across all warehouses."""
	if not batch_item_pairs:
		return {}

	from erpnext.stock.doctype.batch.batch import get_batch_qty

	result = {}
	for item_code, batch_no in batch_item_pairs:
		if not batch_no:
			continue
		try:
			batches = get_batch_qty(
				batch_no=batch_no,
				item_code=item_code,
				for_stock_levels=True,
				consider_negative_batches=True
			)
			total = sum(b.get("qty", 0) for b in (batches or []))
			result[(item_code, batch_no)] = total
		except Exception:
			frappe.log_error(frappe.get_traceback(), f"_fetch_batch_balance failed for item {item_code}, batch {batch_no}")
			result[(item_code, batch_no)] = 0

	return result


# ---------------------------------------------------------------------------
# Print HTML helpers
# ---------------------------------------------------------------------------

def _fetch_si_header_details(si_names):
	"""Bulk-fetch additional Sales Invoice fields not present in the report columns."""
	if not si_names:
		return {}
	placeholders = ", ".join(["%s"] * len(si_names))
	rows = frappe.db.sql(
		f"""
		SELECT
			name,
			po_no,
			territory,
			remarks,
			customer_address,
			shipping_address_name,
			address_display,
			company_address,
			grand_total,
			rounded_total
		FROM `tabSales Invoice`
		WHERE name IN ({placeholders})
		""",
		tuple(si_names),
		as_dict=True,
	)
	result = {}
	for r in rows:
		# Attempt to get company address display
		company_addr = ""
		if r.get("company_address"):
			try:
				addr_doc = frappe.get_cached_doc("Address", r.company_address)
				company_addr = addr_doc.get("address_line1") or ""
				if addr_doc.get("city"):
					company_addr += f", {addr_doc.city}"
				if addr_doc.get("country"):
					company_addr += f", {addr_doc.country}"
			except Exception:
				pass
		r["company_address_display"] = company_addr
		result[r.name] = r
	return result


def _build_filter_summary(filters):
	"""Return a concise human-readable summary of non-empty filters applied."""
	parts = []
	label_map = [
		("sales_invoice", "Sales Invoice"),
		("customer", "Customer"),
		("item_code", "FG Item"),
		("item_group", "Item Group"),
		("batch_no", "FG Batch"),
		("work_order", "Work Order"),
		("raw_material_item", "RM Item"),
		("purchase_receipt", "Purchase Receipt"),
		("customs_document_no", "Customs Doc No"),
	]
	from_date = filters.get("from_date", "")
	to_date = filters.get("to_date", "")
	if from_date or to_date:
		parts.append(f"Period: {from_date} to {to_date}")
	for key, label in label_map:
		val = filters.get(key)
		if val:
			parts.append(f"{label}: {val}")
	return " \u2502 ".join(parts) if parts else "No additional filters applied"


# ---------------------------------------------------------------------------
# Print HTML (whitelisted method)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_print_html(filters):
	if isinstance(filters, str):
		filters = json.loads(filters)
	filters = frappe._dict(filters or {})

	try:
		validate_filters(filters)
	except frappe.ValidationError as e:
		frappe.throw(str(e))

	data = get_data(filters)

	company = frappe.utils.escape_html(filters.get("company", ""))
	print_datetime = frappe.utils.escape_html(str(now_datetime()))
	try:
		printed_by = frappe.utils.escape_html(
			frappe.utils.get_fullname(frappe.session.user) or frappe.session.user
		)
	except Exception:
		printed_by = frappe.utils.escape_html(frappe.session.user or "")
	filter_summary = frappe.utils.escape_html(_build_filter_summary(filters))

	# Group rows by sales invoice, preserving insertion order
	invoices_grouped = {}
	for row in data:
		si = row.get("sales_invoice") or "\u2014"
		invoices_grouped.setdefault(si, []).append(row)

	# Fetch additional SI header details
	real_si_names = [k for k in invoices_grouped if k != "\u2014"]
	si_details = _fetch_si_header_details(real_si_names)

	def _v(val):
		if val is None:
			return ""
		return frappe.utils.escape_html(str(val))

	# Build structured context for the template
	invoices_list = []
	for si_name, rows in invoices_grouped.items():
		first = rows[0]
		si_extra = si_details.get(si_name, {})

		# Build deduplicated FG items list for the header sub-table
		seen_fg_keys = set()
		fg_items = []
		for row in rows:
			fg_key = (
				row.get("si_item_idx"),
				row.get("fg_item_code"),
				row.get("fg_batch_no"),
				row.get("work_order"),
			)
			if fg_key not in seen_fg_keys:
				seen_fg_keys.add(fg_key)
				fg_items.append({
					"si_item_idx": _v(row.get("si_item_idx")),
					"fg_item_code": _v(row.get("fg_item_code")),
					"fg_item_name": _v(row.get("fg_item_name")),
					"sales_qty": _v(row.get("sales_qty")),
					"sales_uom": _v(row.get("sales_uom")),
					"fg_batch_no": _v(row.get("fg_batch_no")),
					"work_order": _v(row.get("work_order")),
					"manufacturing_date": _v(row.get("manufacturing_date")),
				})

		# Build rows (RM / Purchase / Customs only — FG data moved to header)
		row_list = []
		for row in rows:
			row_list.append({
				"rm_item_code": _v(row.get("rm_item_code")),
				"rm_item_name": _v(row.get("rm_item_name")),
				"consumed_qty": _v(row.get("consumed_qty")),
				"rm_batch_no": _v(row.get("rm_batch_no")),
				"purchase_receipt": _v(row.get("purchase_receipt")),
				"purchase_receipt_date": _v(row.get("purchase_receipt_date")),
				"supplier_name": _v(row.get("supplier_name")),
				"pr_qty": _v(row.get("pr_qty")),
				"balance_stock": _v(row.get("balance_stock")),
				"customs_document_no": _v(row.get("customs_document_no")),
			})

		invoices_list.append({
			"si_name": _v(si_name),
			"posting_date": _v(first.get("posting_date")),
			"customer": _v(first.get("customer")),
			"customer_name": _v(first.get("customer_name")),
			"currency": _v(first.get("currency")),
			"company": _v(first.get("company")),
			"po_no": _v(si_extra.get("po_no")),
			"territory": _v(si_extra.get("territory")),
			"remarks": _v(si_extra.get("remarks")),
			"customer_address": _v(si_extra.get("address_display") or si_extra.get("customer_address")),
			"company_address": _v(si_extra.get("company_address_display")),
			"total_amount": _v(si_extra.get("rounded_total") or si_extra.get("grand_total")),
			"fg_items": fg_items,
			"rows": row_list,
		})

	context = {
		"company": company,
		"print_datetime": print_datetime,
		"printed_by": printed_by,
		"filter_summary": filter_summary,
		"invoices": invoices_list,
	}

	template_path = os.path.join(
		os.path.dirname(__file__),
		"customs_export_traceability_report_print.html",
	)
	try:
		with open(template_path, "r") as f:
			template = f.read()
	except OSError as e:
		frappe.throw(f"Could not load print template: {e}")

	return frappe.render_template(template, context)


# ---------------------------------------------------------------------------
# Excel Export (whitelisted method)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_export_excel(filters):
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Alignment, Font, PatternFill
		from openpyxl.utils import get_column_letter
	except ImportError:
		frappe.throw("openpyxl is required to export Excel files. Please install it.")

	if isinstance(filters, str):
		filters = json.loads(filters)
	filters = frappe._dict(filters or {})

	try:
		validate_filters(filters)
	except frappe.ValidationError as e:
		frappe.throw(str(e))

	data = get_data(filters)

	company = filters.get("company", "")
	export_datetime = str(now_datetime())
	try:
		exported_by = frappe.utils.get_fullname(frappe.session.user) or frappe.session.user
	except Exception:
		exported_by = frappe.session.user or ""
	filter_summary = _build_filter_summary(filters)

	# Group rows by sales invoice, preserving insertion order
	invoices_grouped = {}
	for row in data:
		si = row.get("sales_invoice") or "\u2014"
		invoices_grouped.setdefault(si, []).append(row)

	# Fetch additional SI header details
	real_si_names = [k for k in invoices_grouped if k != "\u2014"]
	si_details = _fetch_si_header_details(real_si_names)

	# ---------------------------------------------------------------------------
	# Style helpers
	# ---------------------------------------------------------------------------
	def _make_fill(hex_color):
		return PatternFill(fill_type="solid", fgColor=hex_color)

	def _bold_font(color="000000", size=10):
		return Font(bold=True, color=color, size=size)

	def _normal_font(size=10):
		return Font(size=size)

	ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
	ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
	ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

	# FG table: 8 cols; RM table: 10 cols → use 10 as total width
	TOTAL_COLS = 10

	wb = Workbook()
	ws = wb.active
	ws.title = "Traceability Report"

	# ---------------------------------------------------------------------------
	# Helper: write a merged row spanning all columns
	# ---------------------------------------------------------------------------
	def _write_merged_row(text, font=None, fill=None, align=None):
		row_idx = ws.max_row + 1
		ws.append([""] * TOTAL_COLS)
		cell = ws.cell(row=row_idx, column=1)
		cell.value = text
		if font:
			cell.font = font
		if fill:
			cell.fill = fill
		cell.alignment = align or ALIGN_LEFT
		ws.merge_cells(
			start_row=row_idx, start_column=1,
			end_row=row_idx, end_column=TOTAL_COLS,
		)
		return row_idx

	def _write_blank_row():
		ws.append([""] * TOTAL_COLS)

	# ---------------------------------------------------------------------------
	# A) Report header (written once)
	# ---------------------------------------------------------------------------
	_write_merged_row(
		"Customs Export Traceability Report",
		font=_bold_font(size=14),
		align=ALIGN_CENTER,
	)
	_write_merged_row(company, font=_bold_font(size=11), align=ALIGN_CENTER)
	_write_merged_row(
		f"Exported: {export_datetime} | Exported by: {exported_by}",
		font=_normal_font(size=9),
		align=ALIGN_CENTER,
	)
	_write_merged_row(
		f"Filters applied: {filter_summary}",
		font=_normal_font(size=9),
		align=ALIGN_CENTER,
	)
	_write_blank_row()

	# ---------------------------------------------------------------------------
	# B) Per-invoice sections
	# ---------------------------------------------------------------------------
	FILL_INV_HEADER = _make_fill("EEF3F8")
	FILL_FG_HEADER = _make_fill("2C5F8A")
	FILL_RM_GROUP = _make_fill("3A7A4A")
	FILL_PR_GROUP = _make_fill("7A5A2A")
	FILL_COL_HEADER = _make_fill("3A6B96")
	WHITE_BOLD = _bold_font(color="FFFFFF")

	FG_HEADERS = ["#", "FG Item Code", "FG Item Name", "Sold Qty", "UOM", "FG Batch No", "Work Order", "Mfg Date"]
	RM_HEADERS = [
		"RM Item Code", "RM Item Name", "Consumed Qty", "RM Batch No",
		"Purchase Receipt", "PR Date", "Supplier Name",
		"PR Qty", "Balance Stock", "Customs Doc No",
	]

	for si_name, rows in invoices_grouped.items():
		first = rows[0]
		si_extra = si_details.get(si_name, {})

		# — Invoice header row —
		posting_date = first.get("posting_date") or ""
		customer_name = first.get("customer_name") or ""
		currency = first.get("currency") or ""
		total_amount = si_extra.get("rounded_total") or si_extra.get("grand_total") or ""
		inv_header_text = (
			f"Sales Invoice: {si_name} | Posting Date: {posting_date} | "
			f"Customer: {customer_name} | Currency: {currency} | Total: {total_amount}"
		)
		_write_merged_row(
			inv_header_text,
			font=_bold_font(size=10),
			fill=FILL_INV_HEADER,
		)

		# Optional extra detail rows
		po_no = si_extra.get("po_no") or ""
		territory = si_extra.get("territory") or ""
		remarks = si_extra.get("remarks") or ""
		if po_no or territory:
			extra_parts = []
			if po_no:
				extra_parts.append(f"PO Ref: {po_no}")
			if territory:
				extra_parts.append(f"Territory: {territory}")
			_write_merged_row(
				" | ".join(extra_parts),
				font=_normal_font(size=9),
				fill=FILL_INV_HEADER,
			)
		if remarks:
			_write_merged_row(
				f"Remarks: {remarks}",
				font=_normal_font(size=9),
				fill=FILL_INV_HEADER,
			)

		# — FG sub-header row —
		fg_header_row_idx = ws.max_row + 1
		ws.append(FG_HEADERS + [""] * (TOTAL_COLS - len(FG_HEADERS)))
		for col_idx in range(1, len(FG_HEADERS) + 1):
			cell = ws.cell(row=fg_header_row_idx, column=col_idx)
			cell.font = WHITE_BOLD
			cell.fill = FILL_FG_HEADER
			cell.alignment = ALIGN_CENTER

		# — FG data rows —
		seen_fg_keys = set()
		for row in rows:
			fg_key = (
				row.get("si_item_idx"),
				row.get("fg_item_code"),
				row.get("fg_batch_no"),
				row.get("work_order"),
			)
			if fg_key in seen_fg_keys:
				continue
			seen_fg_keys.add(fg_key)

			mfg_date = row.get("manufacturing_date")
			sales_qty = row.get("sales_qty")
			try:
				sales_qty = float(sales_qty) if sales_qty is not None else ""
			except (TypeError, ValueError):
				sales_qty = str(sales_qty) if sales_qty is not None else ""

			fg_row = [
				row.get("si_item_idx") or "",
				row.get("fg_item_code") or "",
				row.get("fg_item_name") or "",
				sales_qty,
				row.get("sales_uom") or "",
				row.get("fg_batch_no") or "",
				row.get("work_order") or "",
				mfg_date if mfg_date else "",
			]
			data_row_idx = ws.max_row + 1
			ws.append(fg_row + [""] * (TOTAL_COLS - len(fg_row)))
			# Format date cell
			if mfg_date:
				date_cell = ws.cell(row=data_row_idx, column=8)
				date_cell.number_format = "YYYY-MM-DD"

		_write_blank_row()

		# — RM/Purchase/Customs group header row —
		grp_row_idx = ws.max_row + 1
		ws.append([""] * TOTAL_COLS)
		# "Raw Material" spans cols 1-4
		rm_group_cell = ws.cell(row=grp_row_idx, column=1)
		rm_group_cell.value = "Raw Material"
		rm_group_cell.font = WHITE_BOLD
		rm_group_cell.fill = FILL_RM_GROUP
		rm_group_cell.alignment = ALIGN_CENTER
		ws.merge_cells(start_row=grp_row_idx, start_column=1, end_row=grp_row_idx, end_column=4)
		# "Purchase / Customs" spans cols 5-10
		pr_group_cell = ws.cell(row=grp_row_idx, column=5)
		pr_group_cell.value = "Purchase / Customs"
		pr_group_cell.font = WHITE_BOLD
		pr_group_cell.fill = FILL_PR_GROUP
		pr_group_cell.alignment = ALIGN_CENTER
		ws.merge_cells(start_row=grp_row_idx, start_column=5, end_row=grp_row_idx, end_column=10)

		# — RM column header row —
		rm_col_row_idx = ws.max_row + 1
		ws.append(RM_HEADERS)
		for col_idx in range(1, len(RM_HEADERS) + 1):
			cell = ws.cell(row=rm_col_row_idx, column=col_idx)
			cell.font = WHITE_BOLD
			cell.fill = FILL_COL_HEADER
			cell.alignment = ALIGN_CENTER

		# — RM data rows —
		for row in rows:
			pr_date = row.get("purchase_receipt_date")
			consumed_qty = row.get("consumed_qty")
			pr_qty = row.get("pr_qty")
			balance_stock = row.get("balance_stock")

			try:
				consumed_qty = float(consumed_qty) if consumed_qty is not None else ""
			except (TypeError, ValueError):
				consumed_qty = str(consumed_qty) if consumed_qty is not None else ""
			try:
				pr_qty = float(pr_qty) if pr_qty is not None else ""
			except (TypeError, ValueError):
				pr_qty = str(pr_qty) if pr_qty is not None else ""
			try:
				balance_stock = float(balance_stock) if balance_stock is not None else ""
			except (TypeError, ValueError):
				balance_stock = str(balance_stock) if balance_stock is not None else ""

			rm_row = [
				row.get("rm_item_code") or "",
				row.get("rm_item_name") or "",
				consumed_qty,
				row.get("rm_batch_no") or "",
				row.get("purchase_receipt") or "",
				pr_date if pr_date else "",
				row.get("supplier_name") or "",
				pr_qty,
				balance_stock,
				row.get("customs_document_no") or "",
			]
			data_row_idx = ws.max_row + 1
			ws.append(rm_row)
			# Format date cell (column 6)
			if pr_date:
				date_cell = ws.cell(row=data_row_idx, column=6)
				date_cell.number_format = "YYYY-MM-DD"

		_write_blank_row()

	# ---------------------------------------------------------------------------
	# C) Footer
	# ---------------------------------------------------------------------------
	_write_merged_row(
		"\u2014 End of Report \u2014",
		font=_bold_font(size=10),
		align=ALIGN_CENTER,
	)
	_write_merged_row(
		"Blank fields indicate that traceability could not be established from available ERPNext data. "
		"Only submitted documents (Sales Invoice, Stock Entry, Purchase Receipt) are included. "
		"Raw material consumption is based on actual Stock Entry records, not BOM explosion.",
		font=_normal_font(size=8),
		align=ALIGN_CENTER,
	)

	# ---------------------------------------------------------------------------
	# Column widths (reasonable fixed widths)
	# ---------------------------------------------------------------------------
	col_widths = [20, 30, 35, 12, 12, 22, 12, 20, 22, 25]
	for i, width in enumerate(col_widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = width

	# Freeze first 5 header rows
	ws.freeze_panes = "A6"

	# ---------------------------------------------------------------------------
	# Serialize to base64
	# ---------------------------------------------------------------------------
	buf = BytesIO()
	wb.save(buf)
	buf.seek(0)
	file_content = base64.b64encode(buf.read()).decode("utf-8")

	timestamp = now_datetime().strftime("%Y%m%d_%H%M%S")
	file_name = f"Customs_Export_Traceability_Report_{timestamp}.xlsx"

	return {"file_content": file_content, "file_name": file_name}
